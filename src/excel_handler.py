import pandas as pd
from openpyxl import load_workbook

def load_excel_data(file_path):
    return pd.read_excel(file_path)

def update_excel(file_path, item, price, date_col):
    wb = load_workbook(file_path)
    sheet = wb.active
    target_row = None

    # Trouver la rangée correspondant à l'article
    for row in range(1, sheet.max_row + 1):
        if sheet[f"A{row}"].value and sheet[f"A{row}"].value.lower() == item.lower():
            target_row = row
            break

    if target_row is not None:
        # Trouver la colonne pour la date spécifiée
        col = None
        for col_index in range(2, sheet.max_column + 1):
            if sheet.cell(row=1, column=col_index).value == date_col:
                col = col_index
                break

        # Mettre à jour la valeur du prix
        if col:
            sheet.cell(row=target_row, column=col, value=price)
            wb.save(file_path)
            print(f"Updated {item} price to {price} in column {date_col}.")
        else:
            print(f"No column found for date {date_col}.")
    else:
        print(f"No row found for item {item}.")

def add_data_to_excel(file_path, data):
    df = pd.DataFrame(data)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False)
